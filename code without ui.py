# ============================================================
# AI Document Extractor - Local PC Version
# Converted from Google Colab version (no logic changed)
# ============================================================

print("Loading modules... please wait\n")

import google.generativeai as genai
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
import re
from typing import List, Dict
import io
import time
from difflib import SequenceMatcher
from tkinter import Tk, filedialog

# ------------------------------------------------------------
# CONFIGURATION (same as original)
# ------------------------------------------------------------
class Config:
    def __init__(self, api_key: str):
        self.api_key = api_key
        genai.configure(api_key=api_key)

        print("\nSetting up Gemini Flash model...")

        try:
            flash_models = [
                'gemini-2.5-flash',
                'gemini-2.0-flash',
                'gemini-flash-latest',
                'gemini-2.0-flash-exp'
            ]
            model_name = flash_models[0]
            print(f"Using model: {model_name}")
            self.model = genai.GenerativeModel(model_name)

        except:
            print("Model fallback activated")
            self.model = genai.GenerativeModel("gemini-2.5-flash")

# ------------------------------------------------------------
# PDF Extractor (same as original)
# ------------------------------------------------------------
class PDFExtractor:
    def extract_text(self, pdf_file) -> str:
        with pdfplumber.open(pdf_file) as pdf:
            text_content = []
            for page in pdf.pages:
                page_text = page.extract_text(layout=True, x_tolerance=3, y_tolerance=3)
                if page_text:
                    text_content.append(page_text)
        return "\n\n".join(text_content).strip()

# ------------------------------------------------------------
# Comment Optimizer (same as original Colab version)
# ------------------------------------------------------------
class CommentOptimizer:
    @staticmethod
    def similarity_ratio(a, b):
        return SequenceMatcher(None, a.lower(), b.lower()).ratio()

    @staticmethod
    def optimize_comments(data: List[Dict]):
        optimized = []
        seen_comments = {}

        for idx, item in enumerate(data):
            key = item.get('key', "")
            value = item.get('value', "")
            comment = item.get('comments', "").strip()

            # Remove redundant
            if comment and value:
                c_words = set(comment.lower().split())
                v_words = set(value.lower().split())
                k_words = set(key.lower().split())
                new_words = c_words - v_words - k_words
                if len(new_words) < 3:
                    item['comments'] = ""

            # Remove duplicates
            if comment:
                duplicate = False
                for prev, prev_idx in seen_comments.items():
                    if CommentOptimizer.similarity_ratio(comment, prev) > 0.80:
                        duplicate = True
                        item["comments"] = f"See row #{prev_idx + 1}"
                        break

                if not duplicate and len(comment) > 20:
                    seen_comments[comment] = idx

            optimized.append(item)

        return optimized

# ------------------------------------------------------------
# AI Agent (same logic)
# ------------------------------------------------------------
class AIAgent:
    def __init__(self, model):
        self.model = model

    # ---------------------------
    def extract_key_values(self, text):
        prompt = self._prompt(text)
        best_data = []
        best_cov = 0

        for attempt in range(3):
            print(f"\nAI Attempt {attempt + 1}/3...")
            try:
                response = self.model.generate_content(prompt)
                json_data = self._parse_json(response.text)
                
                # Normalize data to ensure keys exist
                valid_data = []
                for i, item in enumerate(json_data, 1):
                    if isinstance(item, dict):
                        norm_item = {
                            "key": str(item.get("key", item.get("Key", ""))),
                            "value": str(item.get("value", item.get("Value", ""))),
                            "comments": str(item.get("comments", item.get("Comments", ""))),
                            "#": i
                        }
                        valid_data.append(norm_item)
                
                json_data = valid_data

                coverage = self.coverage(text, json_data)
                print(f"Coverage: {coverage:.1f}%")

                if coverage > best_cov:
                    best_cov = coverage
                    best_data = json_data

                if coverage >= 98:
                    break
            except Exception as e:
                print(f"Attempt failed: {e}")
                time.sleep(1)

        return best_data

    # ---------------------------
    def _prompt(self, text):
        return f"""
Extract all information from the text below into a JSON array.
Each item must have: "key", "value", "comments".

Rules:
1. Extract every fact, date, and number.
2. Use exact wording for "value".
3. "comments" should be the full sentence context.

Text:
{text}
"""

    # ---------------------------
    def _parse_json(self, txt):
        clean = re.sub(r"```json|```", "", txt)
        m = re.search(r"\[.*\]", clean, re.DOTALL)
        if not m:
            return []
        
        json_str = m.group(0)
        try:
            return json.loads(json_str)
        except json.JSONDecodeError as e:
            if "Extra data" in e.msg:
                try:
                    return json.loads(json_str[:e.pos])
                except:
                    pass
            print(f"JSON Parse Error: {e}")
            return []

    # ---------------------------
    def coverage(self, orig, extracted):
        orig_words = set(re.findall(r"\w+", orig.lower()))
        ext_words = set(re.findall(r"\w+", " ".join(
            e["value"] + " " + e["comments"] for e in extracted
        ).lower()))
        return len(orig_words & ext_words) / len(orig_words) * 100 if orig_words else 0

# ------------------------------------------------------------
# Excel Generator (same)
# ------------------------------------------------------------
class ExcelGenerator:
    def create_excel(self, data):
        df = pd.DataFrame(data)
        base_name = "Output.xlsx"
        output_name = base_name
        
        # Try to find a writable filename
        counter = 1
        while True:
            try:
                with open(output_name, "a"): pass # Check writability
                break
            except PermissionError:
                output_name = f"Output_{counter}.xlsx"
                counter += 1

        df.to_excel(output_name, index=False, sheet_name="Extracted Data")

        wb = load_workbook(output_name)
        ws = wb.active

        header_fill = PatternFill("solid", fgColor="366092")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 80

        ws.freeze_panes = "A2"
        wb.save(output_name)

        return output_name

# ==========================================================
# MAIN FUNCTION (Local version)
# ==========================================================
def main():
    print("============================================")
    print("AI DOCUMENT EXTRACTION SYSTEM - LOCAL MODE")
    print("============================================")

    api_key = input("\nEnter your Gemini API Key: ").strip()
    if not api_key:
        print("Error: API key required")
        return

    print("\nEnter full path to your PDF file:")
    file_path = input("PDF Path: ").strip()

    if not file_path or not file_path.lower().endswith(".pdf"):
        print("❌ Invalid file path or not a PDF.")
        return


    print(f"\nFile selected: {file_path}")

    # Load components
    config = Config(api_key)
    pdf_extractor = PDFExtractor()
    ai_agent = AIAgent(config.model)
    comment_optimizer = CommentOptimizer()
    excel_generator = ExcelGenerator()

    # Extract PDF text
    print("\nExtracting text...")
    raw_text = pdf_extractor.extract_text(file_path)
    print(f"PDF text length: {len(raw_text)} characters")

    # AI Extraction
    print("\nRunning AI extraction...")
    structured = ai_agent.extract_key_values(raw_text)

    # Optimize comments
    structured = comment_optimizer.optimize_comments(structured)

    # Save Excel
    print("\nSaving Excel file...")
    output = excel_generator.create_excel(structured)

    print(f"\nDONE! Excel saved as: {output}\n")

# Run
if __name__ == "__main__":
    main()
