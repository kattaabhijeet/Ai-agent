import google.generativeai as genai
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
import re
from typing import List, Dict
import io
import time
import os

# ------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------
class Config:
    def __init__(self, api_key: str):
        self.api_key = api_key
        genai.configure(api_key=api_key)
        
        # Try to use the best available model
        try:
            flash_models = [
                'gemini-2.5-flash',
                'gemini-2.0-flash',
                'gemini-flash-latest',
                'gemini-2.0-flash-exp'
            ]
            self.model = genai.GenerativeModel(flash_models[0])
        except:
            self.model = genai.GenerativeModel("gemini-2.5-flash")

# ------------------------------------------------------------
# PDF EXTRACTOR
# ------------------------------------------------------------
class PDFExtractor:
    def extract_text(self, pdf_path_or_file) -> str:
        # Handle both file path (str) and file-like object (bytes)
        with pdfplumber.open(pdf_path_or_file) as pdf:
            text_content = []
            for page in pdf.pages:
                page_text = page.extract_text(layout=True, x_tolerance=3, y_tolerance=3)
                if page_text:
                    text_content.append(page_text)
        return "\n\n".join(text_content).strip()

# ------------------------------------------------------------
# COMMENT OPTIMIZER
# ------------------------------------------------------------
from difflib import SequenceMatcher

class CommentOptimizer:
    @staticmethod
    def similarity_ratio(a, b):
        return SequenceMatcher(None, a.lower(), b.lower()).ratio()

    @staticmethod
    def optimize_comments(data: List[Dict]):
        optimized = []
        seen_comments = {}

        for idx, item in enumerate(data):
            key = item.get('key', "")
            value = item.get('value', "")
            comment = item.get('comments', "").strip()

            # Skip redundancy removal - keep all comments as-is
            # The AI should provide meaningful comments
            
            # Remove duplicates - leave them empty instead of showing "See row #X"
            if comment and len(comment) > 10:  # Only process substantial comments
                duplicate = False
                for prev, prev_idx in seen_comments.items():
                    if CommentOptimizer.similarity_ratio(comment, prev) > 0.85:
                        duplicate = True
                        item["comments"] = ""  # Leave empty instead of "See row #X"
                        break

                if not duplicate:
                    seen_comments[comment] = idx

            optimized.append(item)

        return optimized

# ------------------------------------------------------------
# AI AGENT
# ------------------------------------------------------------
class AIAgent:
    def __init__(self, model):
        self.model = model
    
    def extract_key_values(self, text: str) -> List[Dict]:
        prompt = f"""Extract ALL information from this text.
        
IMPORTANT INSTRUCTIONS:
1. Extract 100% of the content - every fact, number, date, name.
2. Use EXACT original wording in the "value" field.
3. Create separate entries for EACH piece of information.
4. For the "comments" field: Capture the FULL sentence or phrase from the text that provides context.
5. Output format: Use a pipe (|) separated list. One entry per line.
   Format: Key | Value | Comments
   
   Example:
   Name | Vijay Kumar | Vijay Kumar was born on March 15, 1989.
   Age | 35 | He is 35 years old as of 2024.

TEXT TO EXTRACT:
{text}"""
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                response = self.model.generate_content(
                    prompt,
                    generation_config={'temperature': 0.1, 'max_output_tokens': 8192}
                )
                print(f"DEBUG: Raw AI Response: {response.text[:500]}...")
                return self._extract_pipe_separated(response.text)
            except Exception as e:
                print(f"DEBUG: Extraction Error: {e}")
                if attempt < max_retries - 1: time.sleep(2)
        
        raise Exception("Extraction failed after retries")
    
    def _extract_pipe_separated(self, text: str) -> List[Dict]:
        valid_data = []
        lines = text.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('Key |') or line.startswith('---'):
                continue
                
            parts = line.split('|')
            if len(parts) >= 2:
                key = parts[0].strip()
                value = parts[1].strip()
                comments = parts[2].strip() if len(parts) > 2 else ""
                
                if key or value:
                    valid_data.append({
                        'key': key,
                        'value': value,
                        'comments': comments
                    })
        
        print(f"DEBUG: Parsed {len(valid_data)} items from pipe-separated text")
        return valid_data

    def _extract_json(self, text: str) -> List[Dict]:
        # Deprecated but kept for compatibility if needed
        return []

# ------------------------------------------------------------
# EXCEL GENERATOR
# ------------------------------------------------------------
class ExcelGenerator:
    def create_excel(self, data: List[Dict], output_path: str):
        df = pd.DataFrame(data)
        df = df.rename(columns={'#': '#', 'key': 'Key', 'value': 'Value', 'comments': 'Comments'})
        
        columns = ['#', 'Key', 'Value', 'Comments']
        for col in columns:
            if col not in df.columns: df[col] = ''
        
        df = df[columns]
        df.to_excel(output_path, index=False, sheet_name='Extracted Data')
        self._apply_formatting(output_path)
        return output_path
    
    def _apply_formatting(self, excel_path: str):
        wb = load_workbook(excel_path)
        ws = wb.active
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 80
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        ws.freeze_panes = 'A2'
        wb.save(excel_path)
